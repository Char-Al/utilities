#!/usr/bin/env python3

__author__ = 'Charles Van Goethem'
__license__ = 'unlicense'
__version__ = '0.0.1'
__email__ = 'c-vangoethem@chu-montpellier.fr'
__status__ = 'dev'

import openpyxl
import argh
import tqdm


@argh.arg(
    'files', metavar="fileN.xlsx",
    help="List of files to merge with file 1",
    nargs='+', type=str)
@argh.arg('--output', help="Output file to save.")
def merge_files(files, output="merge.xlsx"):
    "Merge ultiple xlsx files in one. Each sheet dispatch in a file."
    wb_final = openpyxl.Workbook()

    for sheet in wb_final.sheetnames:
        std = wb_final[sheet]
        wb_final.remove(std)

    for f in tqdm.tqdm(files, total=len(files)):
        wb_onload = openpyxl.load_workbook(f)
        for sheet in wb_onload:
            ws = wb_final.create_sheet(sheet.title)
            for row in sheet:
                for cell in row:
                    ws[cell.coordinate].value = cell.value
                    ws[cell.coordinate].style = cell.style

    wb_final.save(output)


@argh.arg(
    'file', metavar="file.xlsx", type=str,
    help="File wich header will be removed (default: active sheet/1st line).")
@argh.arg(
    '--sheets', metavar="name",
    help="List of sheet names wich header will be removed.",
    nargs='*', type=str)
@argh.arg('--all-sheets', help="Remove header to all sheets.")
@argh.arg(
    '--lines',
    help="Number of row of the header.", nargs='+', type=int)
@argh.arg('--output', help="Output file (default same file).")
def remove_lines(file, sheets=None, all_sheets=False, lines=[1], output=None):
    "Remove line(s) from sheet(s)."
    workbook = openpyxl.load_workbook(file)

    list_of_sheets = [workbook.active.title]
    if sheets is not None:
        list_of_sheets = sheets
    if all_sheets:
        list_of_sheets = workbook.sheetnames

    for sheet in list_of_sheets:
        for line in lines:
            workbook[sheet].delete_rows(line)

    out_file = str()
    out_file = file if output is None else output
    workbook.save(out_file)


@argh.arg(
    'file', metavar="file.xlsx", type=str,
    help="File to concatenate.")
@argh.arg('--no-header', help="There is no header in sheets.")
@argh.arg('--out-sheet', help="Write in a new sheet with this name.")
@argh.arg('--output', help="Output file (default same file).")
def concatenate_sheets(file, no_header=True, out_sheet="concat", output=None):
    "Concatenate all sheet(s) into 1 sheet."
    workbook = openpyxl.load_workbook(file)

    ws = workbook.create_sheet(out_sheet)
    line = 0
    for sheet in workbook.sheetnames:
        if sheet != out_sheet:
            print(sheet)
            for current_line, row in enumerate(workbook[sheet].rows, start=1):
                if (
                    (no_header is False and line == 0 and current_line == 1) or
                    (no_header is False and current_line > 1) or
                    (no_header is True)
                ):
                    line += 1
                    for column, cell in enumerate(row, start=1):
                        column_letter = openpyxl.utils.get_column_letter(column)
                        ws[f"{column_letter}{line}"].value = cell.value

    out_file = str()
    out_file = file if output is None else output
    workbook.save(out_file)


if __name__ == '__main__':
    parser = argh.ArghParser()
    parser.add_commands([merge_files, remove_lines, concatenate_sheets])
    parser.dispatch()
